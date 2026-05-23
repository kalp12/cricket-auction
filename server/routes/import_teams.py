import io
import csv
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

from db.database import get_db
from auth.auth import get_current_user
from models.models import Team, Auction

router = APIRouter()

TEAM_FIELDS = {
    "name": "Team Name",
    "short_name": "Short Name",
    "total_budget": "Total Budget",
    "max_players": "Max Players",
    "logo_url": "Logo URL",
}

REQUIRED_FIELDS = {"name", "total_budget"}

TEMPLATE_HEADERS = [
    "Team Name", "Short Name", "Total Budget", "Max Players", "Logo URL",
]

TEMPLATE_FIELD_MAP = {
    "Team Name": "name",
    "Short Name": "short_name",
    "Total Budget": "total_budget",
    "Max Players": "max_players",
    "Logo URL": "logo_url",
}

COMMON_ALIASES = {
    "team": "name", "team name": "name", "team_name": "name", "name": "name",
    "short": "short_name", "short name": "short_name", "short_name": "short_name", "abbr": "short_name", "abbreviation": "short_name",
    "budget": "total_budget", "total budget": "total_budget", "total_budget": "total_budget", "purse": "total_budget",
    "max": "max_players", "max players": "max_players", "max_players": "max_players", "squad size": "max_players",
    "logo": "logo_url", "logo url": "logo_url", "logo_url": "logo_url",
}


@router.get("/teams/template")
async def download_team_template(
    auction_id: int = Query(...),
    current_user: dict = Depends(get_current_user),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Teams"
    ws.append(TEMPLATE_HEADERS)
    # Sample rows
    ws.append(["Mumbai Indians", "MI", 100000000, 25, ""])
    ws.append(["Chennai Super Kings", "CSK", 100000000, 25, ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=team_import_template.xlsx"},
    )


@router.post("/teams/upload")
async def upload_team_file(
    auction_id: int = Query(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    auction = db.query(Auction).filter(Auction.id == auction_id).first()
    if not auction:
        raise HTTPException(404, "Auction not found")

    contents = await file.read()
    filename = file.filename or ""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    headers = []
    rows = []
    try:
        if ext == "xlsx":
            wb = load_workbook(io.BytesIO(contents), read_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                raise HTTPException(400, "File is empty")
            headers = [str(h or "").strip() for h in all_rows[0]]
            for row in all_rows[1:]:
                rows.append([str(v or "").strip() if v is not None else "" for v in row])
        elif ext == "csv":
            text = contents.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(text))
            all_rows = list(reader)
            if not all_rows:
                raise HTTPException(400, "File is empty")
            headers = [str(h or "").strip() for h in all_rows[0]]
            for row in all_rows[1:]:
                rows.append([str(v or "").strip() for v in row])
        else:
            raise HTTPException(400, "Unsupported file type. Use .xlsx or .csv")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Failed to parse file: {str(e)}")

    # Auto-suggest column mapping
    suggested_mapping = {}
    for idx, header in enumerate(headers):
        key = header.lower().strip()
        if key in COMMON_ALIASES:
            suggested_mapping[str(idx)] = COMMON_ALIASES[key]

    return {
        "headers": headers,
        "rows": rows[:50],
        "total_rows": len(rows),
        "suggested_mapping": suggested_mapping,
        "available_fields": TEAM_FIELDS,
    }


@router.post("/teams/commit")
async def commit_team_import(
    data: dict,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    auction_id = data.get("auction_id")
    mapping = data.get("mapping", {})
    rows = data.get("rows", [])

    if not auction_id:
        raise HTTPException(400, "auction_id required")

    auction = db.query(Auction).filter(Auction.id == auction_id).first()
    if not auction:
        raise HTTPException(404, "Auction not found")

    # Build reverse mapping: field_key -> col_idx
    field_to_col = {}
    for col_idx, field_key in mapping.items():
        if field_key in TEAM_FIELDS:
            field_to_col[field_key] = int(col_idx)

    teams_created = 0
    errors = []

    try:
        for row_idx, row in enumerate(rows):
            try:
                def get_val(key):
                    col = field_to_col.get(key)
                    if col is None or col >= len(row):
                        return None
                    v = row[col].strip()
                    return v if v else None

                name = get_val("name")
                if not name:
                    errors.append(f"Row {row_idx + 1}: team name is required")
                    continue

                budget_str = get_val("total_budget")
                if not budget_str:
                    errors.append(f"Row {row_idx + 1}: total budget is required")
                    continue

                try:
                    budget = float(budget_str.replace(",", ""))
                except ValueError:
                    errors.append(f"Row {row_idx + 1}: invalid budget value '{budget_str}'")
                    continue

                max_players = None
                mp_str = get_val("max_players")
                if mp_str:
                    try:
                        max_players = int(float(mp_str.replace(",", "")))
                    except ValueError:
                        pass

                team = Team(
                    auction_id=auction_id,
                    name=name,
                    short_name=get_val("short_name") or name[:3].upper(),
                    total_budget=budget,
                    remaining_budget=budget,
                    max_players=max_players,
                    logo_url=get_val("logo_url"),
                )
                db.add(team)
                teams_created += 1

            except Exception as e:
                errors.append(f"Row {row_idx + 1}: {str(e)}")

        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Team import failed: {str(e)}")

    return {
        "teams_created": teams_created,
        "errors": errors[:20],
        "total_rows": len(rows),
    }
