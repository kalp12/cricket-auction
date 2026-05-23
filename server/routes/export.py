import io
import csv
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from db.database import get_db
from auth.auth import get_current_user
from models.models import Player, Team, TeamPlayer, Auction, Bid

router = APIRouter()

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)


def style_header(ws, headers, row=1):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws):
    for col in ws.columns:
        first = col[0]
        if hasattr(first, 'column_letter'):
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[first.column_letter].width = min(max_length + 4, 40)


def format_price(val):
    if val >= 10000000:
        return f"₹{val/10000000:.1f} Cr"
    if val >= 100000:
        return f"₹{val/100000:.1f} L"
    return f"₹{val:,.0f}"


# ── Export Players ──────────────────────────────────────
@router.get("/players")
async def export_players(
    auction_id: int = Query(...),
    format: str = Query("xlsx"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    auction = db.query(Auction).filter(Auction.id == auction_id).first()
    if not auction:
        raise HTTPException(404, "Auction not found")

    players = db.query(Player).filter(Player.auction_id == auction_id).order_by(Player.name).all()
    headers = ["Name", "Role", "Country", "Base Price", "Status", "Matches", "Runs", "Wickets", "Batting Avg", "Batting SR", "Bowling Avg", "Bowling Econ", "Image URL"]

    rows_data = []
    for p in players:
        rows_data.append([
            p.name, p.role, p.country, p.base_price, p.status,
            p.matches or 0, p.runs or 0, p.wickets or 0,
            p.batting_avg or 0, p.batting_sr or 0, p.bowling_avg or 0, p.bowling_econ or 0,
            p.image_url or "",
        ])

    filename = f"{auction.name.replace(' ', '_')}_players"

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows_data)
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
        )

    # xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = "Players"
    style_header(ws, headers)
    for row_idx, row in enumerate(rows_data, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val).border = THIN_BORDER
    auto_width(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )


# ── Export Auction Results ──────────────────────────────
@router.get("/auction-results")
async def export_auction_results(
    auction_id: int = Query(...),
    format: str = Query("xlsx"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    auction = db.query(Auction).filter(Auction.id == auction_id).first()
    if not auction:
        raise HTTPException(404, "Auction not found")

    players = db.query(Player).filter(Player.auction_id == auction_id).order_by(Player.name).all()
    team_ids = [t.id for t in db.query(Team).filter(Team.auction_id == auction_id).all()]
    team_players = db.query(TeamPlayer).filter(TeamPlayer.team_id.in_(team_ids)).all()
    tp_map = {tp.player_id: (tp.team_id, tp.bought_price) for tp in team_players}
    teams = db.query(Team).filter(Team.auction_id == auction_id).all()
    team_map = {t.id: t.name for t in teams}

    headers = ["Name", "Role", "Country", "Base Price", "Status", "Sold Price", "Team", "Matches", "Runs", "Wickets", "Batting Avg", "Batting SR", "Bowling Avg", "Bowling Econ"]

    rows_data = []
    for p in players:
        sold_price = ""
        team_name = ""
        if p.status == "sold" and p.id in tp_map:
            _, bp = tp_map[p.id]
            sold_price = bp
            tid = tp_map[p.id][0]
            team_name = team_map.get(tid, "Unknown")
        rows_data.append([
            p.name, p.role, p.country, p.base_price, p.status,
            sold_price, team_name,
            p.matches or 0, p.runs or 0, p.wickets or 0,
            p.batting_avg or 0, p.batting_sr or 0, p.bowling_avg or 0, p.bowling_econ or 0,
        ])

    filename = f"{auction.name.replace(' ', '_')}_results"

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows_data)
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Auction Results"
    style_header(ws, headers)
    for row_idx, row in enumerate(rows_data, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            # Highlight sold rows in green
            if row[4] == "sold":
                cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            elif row[4] == "unsold":
                cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    auto_width(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )


# ── Export Team Rosters ─────────────────────────────────
@router.get("/team-rosters")
async def export_team_rosters(
    auction_id: int = Query(...),
    format: str = Query("xlsx"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    auction = db.query(Auction).filter(Auction.id == auction_id).first()
    if not auction:
        raise HTTPException(404, "Auction not found")

    teams = db.query(Team).filter(Team.auction_id == auction_id).order_by(Team.name).all()
    team_ids = [t.id for t in teams]
    team_players = db.query(TeamPlayer).filter(TeamPlayer.team_id.in_(team_ids)).all()
    players = db.query(Player).filter(Player.auction_id == auction_id).all()
    player_map = {p.id: p for p in players}

    filename = f"{auction.name.replace(' ', '_')}_team_rosters"

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Team", "Short Name", "Budget", "Remaining", "Player", "Role", "Country", "Bought Price", "Matches", "Runs", "Wickets"])
        for team in teams:
            for tp in team_players:
                if tp.team_id == team.id:
                    p = player_map.get(tp.player_id)
                    if p:
                        writer.writerow([
                            team.name, team.short_name or "", team.total_budget, team.remaining_budget,
                            p.name, p.role, p.country, tp.bought_price,
                            p.matches or 0, p.runs or 0, p.wickets or 0,
                        ])
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
        )

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    TEAM_HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    TEAM_HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)

    for team in teams:
        ws = wb.create_sheet(title=team.name[:31])  # sheet name max 31 chars

        # Team header row
        ws.merge_cells("A1:K1")
        cell = ws.cell(row=1, column=1, value=f"{team.name} ({team.short_name or ''})")
        cell.font = TEAM_HEADER_FONT
        cell.fill = TEAM_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

        # Budget row
        ws.merge_cells("A2:K2")
        budget_cell = ws.cell(row=2, column=1, value=f"Budget: {format_price(team.total_budget)} | Remaining: {format_price(team.remaining_budget)}")
        budget_cell.font = Font(italic=True, color="6B7280", size=10)
        budget_cell.alignment = Alignment(horizontal="center")

        # Player headers
        player_headers = ["#", "Player", "Role", "Country", "Bought Price", "Matches", "Runs", "Wickets", "Bat Avg", "Bat SR", "Bowl Econ"]
        style_header(ws, player_headers, row=4)

        row_idx = 5
        count = 0
        total_spent = 0
        for tp in team_players:
            if tp.team_id == team.id:
                p = player_map.get(tp.player_id)
                if p:
                    count += 1
                    total_spent += tp.bought_price
                    row_data = [count, p.name, p.role, p.country, tp.bought_price, p.matches or 0, p.runs or 0, p.wickets or 0, p.batting_avg or 0, p.batting_sr or 0, p.bowling_econ or 0]
                    for col_idx, val in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=val).border = THIN_BORDER
                    row_idx += 1

        # Summary row
        row_idx += 1
        ws.merge_cells(f"A{row_idx}:D{row_idx}")
        summary_cell = ws.cell(row=row_idx, column=1, value=f"Total: {count} players | Spent: {format_price(total_spent)}")
        summary_cell.font = Font(bold=True, size=10)

        auto_width(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )
