import io
import csv
from difflib import SequenceMatcher
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Query
from sqlalchemy.orm import Session
from openpyxl.reader.excel import load_workbook

from db.database import get_db
from auth.auth import get_current_user
from models.models import Player, Auction, StatUpdate

router = APIRouter()

STAT_FIELDS = {
    "matches": "Matches",
    "runs": "Runs",
    "wickets": "Wickets",
    "batting_avg": "Batting Average",
    "batting_sr": "Batting Strike Rate",
    "bowling_avg": "Bowling Average",
    "bowling_econ": "Bowling Economy",
}

# Map common external column names to our fields
COMMON_ALIASES = {
    "mat": "matches", "match": "matches", "matches": "matches",
    "run": "runs", "runs": "runs", "runs scored": "runs",
    "wkt": "wickets", "wickets": "wickets", "wkt taken": "wickets",
    "bat avg": "batting_avg", "batting average": "batting_avg", "avg": "batting_avg", "batting_avg": "batting_avg",
    "sr": "batting_sr", "strike rate": "batting_sr", "bat sr": "batting_sr", "batting_sr": "batting_sr",
    "bowl avg": "bowling_avg", "bowling average": "bowling_avg", "bowl_avg": "bowling_avg", "bowling_avg": "bowling_avg",
    "econ": "bowling_econ", "economy": "bowling_econ", "econ rate": "bowling_econ", "bowling_econ": "bowling_econ",
    "player": "name", "name": "name", "player name": "name", "player_name": "name",
    "player_id": "player_id", "id": "player_id",
}


def fuzzy_match(name: str, candidates: dict, threshold: float = 0.6) -> int | None:
    """Fuzzy match a name against candidate dict {name: player_id}. Returns player_id or None."""
    name_lower = name.lower().strip()
    # Exact match first
    for cand_name, pid in candidates.items():
        if cand_name.lower() == name_lower:
            return pid
    # Fuzzy match
    best_score = 0.0
    best_pid = None
    for cand_name, pid in candidates.items():
        score = SequenceMatcher(None, name_lower, cand_name.lower()).ratio()
        if score > best_score:
            best_score = score
            best_pid = pid
    if best_score >= threshold:
        return best_pid
    return None


@router.post("/upload")
async def upload_stats_file(
    auction_id: int = Query(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    auction = db.query(Auction).filter(Auction.id == auction_id).first()
    if not auction:
        raise HTTPException(404, "Auction not found")

    contents = await file.read()
    filename = file.filename or ""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    headers = []
    rows = []
    try:
        if ext == "xlsx":
            wb = load_workbook(io.BytesIO(contents), read_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                raise HTTPException(400, "File is empty")
            headers = [str(h or "").strip() for h in all_rows[0]]
            for row in all_rows[1:]:
                rows.append([str(v or "").strip() if v is not None else "" for v in row])
        elif ext == "csv":
            text = contents.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(text))
            all_rows = list(reader)
            if not all_rows:
                raise HTTPException(400, "File is empty")
            headers = [str(h or "").strip() for h in all_rows[0]]
            for row in all_rows[1:]:
                rows.append([str(v or "").strip() for v in row])
        else:
            raise HTTPException(400, "Unsupported file type. Use .xlsx or .csv")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Failed to parse file: {str(e)}")

    # Auto-suggest column mapping using aliases
    suggested_mapping = {}
    for idx, header in enumerate(headers):
        key = header.lower().strip()
        if key in COMMON_ALIASES:
            suggested_mapping[str(idx)] = COMMON_ALIASES[key]

    # Build player lookup for this auction
    players = db.query(Player).filter(Player.auction_id == auction_id).order_by(Player.name).all()
    player_lookup = {p.name: p.id for p in players}

    # Auto-match rows to players (first 50 rows for preview)
    matched_rows = []
    for row in rows[:50]:
        # Try to find player name column
        name_col = None
        pid_col = None
        for col_idx, field_key in suggested_mapping.items():
            if field_key == "name":
                name_col = int(col_idx)
            elif field_key == "player_id":
                pid_col = int(col_idx)
        # Also scan unmapped columns for "name"/"player" patterns
        if name_col is None:
            for idx, header in enumerate(headers):
                if str(idx) not in suggested_mapping:
                    hl = header.lower().strip()
                    if hl in ("player", "name", "player name", "player_name"):
                        name_col = idx
                        break
        player_id = None
        player_name = ""
        confidence = ""
        if pid_col is not None and pid_col < len(row) and row[pid_col].strip().isdigit():
            pid = int(row[pid_col].strip())
            if pid in [p.id for p in players]:
                player_id = pid
                player_name = next(p.name for p in players if p.id == pid)
                confidence = "exact_id"
        if player_id is None and name_col is not None and name_col < len(row):
            raw_name = row[name_col].strip()
            player_name = raw_name
            pid = fuzzy_match(raw_name, player_lookup)
            if pid is not None:
                player_id = pid
                actual = next(p.name for p in players if p.id == pid)
                confidence = "exact" if actual.lower() == raw_name.lower() else "fuzzy"
        matched_rows.append({
            "row": row[:len(headers)],
            "player_id": player_id,
            "player_name": player_name,
            "confidence": confidence,
        })

    return {
        "headers": headers,
        "rows": rows[:50],
        "total_rows": len(rows),
        "suggested_mapping": suggested_mapping,
        "available_stat_fields": STAT_FIELDS,
        "player_names": {p.id: p.name for p in players},
        "matched_rows": matched_rows,
    }


@router.post("/commit")
async def commit_stats_import(
    data: dict,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    auction_id = data.get("auction_id")
    source = data.get("source", "custom")
    mapping = data.get("mapping", {})  # col_idx -> field_key (only stat fields + name/player_id)
    rows = data.get("rows", [])
    player_overrides = data.get("player_overrides", {})  # row_idx -> player_id (manual fixes)

    if not auction_id:
        raise HTTPException(400, "auction_id required")

    # Build reverse: field_key -> col_idx
    field_to_col = {}
    for col_idx, field_key in mapping.items():
        if field_key in STAT_FIELDS or field_key in ("name", "player_id"):
            field_to_col[field_key] = int(col_idx)

    players = db.query(Player).filter(Player.auction_id == auction_id).all()
    player_lookup = {p.name: p.id for p in players}
    player_map = {p.id: p for p in players}

    updates_applied = 0
    errors = []

    try:
        for row_idx, row in enumerate(rows):
            try:
                # Resolve player_id
                player_id = player_overrides.get(str(row_idx))

                if player_id is None and "player_id" in field_to_col:
                    col = field_to_col["player_id"]
                    if col < len(row) and row[col].strip().isdigit():
                        pid = int(row[col].strip())
                        if pid in player_map:
                            player_id = pid

                if player_id is None and "name" in field_to_col:
                    col = field_to_col["name"]
                    if col < len(row) and row[col].strip():
                        pid = fuzzy_match(row[col].strip(), player_lookup)
                        if pid is not None:
                            player_id = pid

                if player_id is None:
                    errors.append(f"Row {row_idx + 1}: could not match to a player")
                    continue

                player = player_map.get(player_id)
                if not player:
                    errors.append(f"Row {row_idx + 1}: player {player_id} not found")
                    continue

                def get_val(field_key: str):
                    col = field_to_col.get(field_key)
                    if col is None or col >= len(row):
                        return None
                    v = row[col].strip()
                    return v if v else None

                def int_val(field_key: str):
                    v = get_val(field_key)
                    if not v:
                        return None
                    try:
                        return int(float(v.replace(",", "")))
                    except (ValueError, TypeError):
                        return None

                def float_val(field_key: str):
                    v = get_val(field_key)
                    if not v:
                        return None
                    try:
                        return float(v.replace(",", ""))
                    except (ValueError, TypeError):
                        return None

                # Check if any stats actually changed
                changes = {}
                int_fields = {"matches", "runs", "wickets"}
                float_fields = {"batting_avg", "batting_sr", "bowling_avg", "bowling_econ"}

                for f in int_fields:
                    if f in field_to_col:
                        new = int_val(f)
                        if new is not None and new != (getattr(player, f) or 0):
                            changes[f] = new

                for f in float_fields:
                    if f in field_to_col:
                        new = float_val(f)
                        if new is not None and abs(new - (getattr(player, f) or 0)) > 0.01:
                            changes[f] = new

                if not changes:
                    continue  # No actual changes

                # Record the stat update for versioning
                stat_update = StatUpdate(
                    player_id=player.id,
                    source=source,
                    old_matches=player.matches if "matches" in changes else None,
                    new_matches=changes.get("matches"),
                    old_runs=player.runs if "runs" in changes else None,
                    new_runs=changes.get("runs"),
                    old_wickets=player.wickets if "wickets" in changes else None,
                    new_wickets=changes.get("wickets"),
                    old_batting_avg=player.batting_avg if "batting_avg" in changes else None,
                    new_batting_avg=changes.get("batting_avg"),
                    old_batting_sr=player.batting_sr if "batting_sr" in changes else None,
                    new_batting_sr=changes.get("batting_sr"),
                    old_bowling_avg=player.bowling_avg if "bowling_avg" in changes else None,
                    new_bowling_avg=changes.get("bowling_avg"),
                    old_bowling_econ=player.bowling_econ if "bowling_econ" in changes else None,
                    new_bowling_econ=changes.get("bowling_econ"),
                )
                db.add(stat_update)

                # Apply the changes to the player
                for field_key, new_val in changes.items():
                    setattr(player, field_key, new_val)

                updates_applied += 1

            except Exception as e:
                errors.append(f"Row {row_idx + 1}: {str(e)}")

        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Stats import failed: {str(e)}")

    return {
        "updates_applied": updates_applied,
        "errors": errors[:20],
        "total_rows": len(rows),
    }


@router.get("/history/{player_id}")
async def get_stats_history(
    player_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    player = db.query(Player).filter(Player.id == player_id).first()
    if not player:
        raise HTTPException(404, "Player not found")

    updates = db.query(StatUpdate).filter(
        StatUpdate.player_id == player_id
    ).order_by(StatUpdate.created_at.desc()).all()

    return {
        "player_id": player_id,
        "player_name": player.name,
        "current_stats": {
            "matches": player.matches,
            "runs": player.runs,
            "wickets": player.wickets,
            "batting_avg": player.batting_avg,
            "batting_sr": player.batting_sr,
            "bowling_avg": player.bowling_avg,
            "bowling_econ": player.bowling_econ,
        },
        "history": [
            {
                "id": u.id,
                "source": u.source,
                "created_at": u.created_at.isoformat() if u.created_at else None,
                "old_matches": u.old_matches,
                "new_matches": u.new_matches,
                "old_runs": u.old_runs,
                "new_runs": u.new_runs,
                "old_wickets": u.old_wickets,
                "new_wickets": u.new_wickets,
                "old_batting_avg": u.old_batting_avg,
                "new_batting_avg": u.new_batting_avg,
                "old_batting_sr": u.old_batting_sr,
                "new_batting_sr": u.new_batting_sr,
                "old_bowling_avg": u.old_bowling_avg,
                "new_bowling_avg": u.new_bowling_avg,
                "old_bowling_econ": u.old_bowling_econ,
                "new_bowling_econ": u.new_bowling_econ,
            }
            for u in updates
        ],
    }
