import io
import csv
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

from db.database import get_db
from auth.auth import get_current_user
from models.models import Player, Auction

router = APIRouter()

PLAYER_FIELDS = {
    "name": "Player Name",
    "role": "Role (batsman/bowler/allrounder/wicketkeeper)",
    "country": "Country",
    "base_price": "Base Price",
    "image_url": "Image URL",
    "matches": "Matches",
    "runs": "Runs",
    "wickets": "Wickets",
    "batting_avg": "Batting Average",
    "batting_sr": "Batting Strike Rate",
    "bowling_avg": "Bowling Average",
    "bowling_econ": "Bowling Economy",
}

REQUIRED_FIELDS = {"name", "role", "country", "base_price"}

TEMPLATE_HEADERS = [
    "Player Name", "Role", "Country", "Base Price",
    "Image URL", "Matches", "Runs", "Wickets",
    "Batting Average", "Batting Strike Rate",
    "Bowling Average", "Bowling Economy",
]

TEMPLATE_FIELD_MAP = {
    "Player Name": "name",
    "Role": "role",
    "Country": "country",
    "Base Price": "base_price",
    "Image URL": "image_url",
    "Matches": "matches",
    "Runs": "runs",
    "Wickets": "wickets",
    "Batting Average": "batting_avg",
    "Batting Strike Rate": "batting_sr",
    "Bowling Average": "bowling_avg",
    "Bowling Economy": "bowling_econ",
}


@router.get("/players/template")
async def download_template(
    auction_id: int = Query(...),
    current_user: dict = Depends(get_current_user),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Players"

    for col_idx, header in enumerate(TEMPLATE_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.style = "Headline 1"

    sample = ["Virat Kohli", "batsman", "India", 20000000, "", 274, 13848, 4, 57.7, 93.4, 0, 0]
    for col_idx, val in enumerate(sample, 1):
        ws.cell(row=2, column=col_idx, value=val)

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=player_import_template.xlsx"},
    )


@router.post("/players/upload")
async def upload_player_file(
    auction_id: int = Query(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    auction_obj = db.query(Auction).filter(Auction.id == auction_id).first()
    if not auction_obj:
        raise HTTPException(status_code=404, detail="Auction not found")

    contents = await file.read()
    filename = file.filename or ""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    headers = []
    rows = []

    try:
        if ext == "xlsx":
            wb = load_workbook(io.BytesIO(contents), read_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                raise HTTPException(status_code=400, detail="File is empty")
            headers = [str(h or "").strip() for h in all_rows[0]]
            for row in all_rows[1:]:
                rows.append([str(v or "").strip() if v is not None else "" for v in row])
        elif ext == "csv":
            text = contents.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(text))
            all_rows = list(reader)
            if not all_rows:
                raise HTTPException(status_code=400, detail="File is empty")
            headers = [str(h or "").strip() for h in all_rows[0]]
            for row in all_rows[1:]:
                rows.append([str(v or "").strip() for v in row])
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type. Use .xlsx or .csv")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    suggested_mapping = {}
    for idx, header in enumerate(headers):
        header_lower = header.lower().strip()
        for template_header, field_key in TEMPLATE_FIELD_MAP.items():
            if header_lower == template_header.lower() or header_lower == field_key:
                suggested_mapping[str(idx)] = field_key
                break

    return {
        "headers": headers,
        "rows": rows[:50],
        "total_rows": len(rows),
        "suggested_mapping": suggested_mapping,
        "available_fields": list(PLAYER_FIELDS.values()),
    }


@router.post("/players/commit")
async def commit_import(
    data: dict,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    auction_id = data.get("auction_id")
    mapping = data.get("mapping", {})
    rows = data.get("rows", [])

    if not auction_id:
        raise HTTPException(status_code=400, detail="auction_id required")

    auction = db.query(Auction).filter(Auction.id == auction_id).first()
    if not auction:
        raise HTTPException(status_code=404, detail="Auction not found")

    # Build field_key → column_index from mapping
    field_to_col = {}
    for col_idx, field_value in mapping.items():
        if field_value in PLAYER_FIELDS:
            field_to_col[field_value] = int(col_idx)
        else:
            # Check by display name
            for field_key, display in PLAYER_FIELDS.items():
                if field_value == display:
                    field_to_col[field_key] = int(col_idx)
                    break

    for req in REQUIRED_FIELDS:
        if req not in field_to_col:
            raise HTTPException(status_code=400, detail=f"Required field '{PLAYER_FIELDS[req]}' is not mapped")

    errors = []
    players_created = 0

    try:
        for row_idx, row in enumerate(rows):
            try:
                def get_val(field_key: str):
                    col = field_to_col.get(field_key)
                    if col is None or col >= len(row):
                        return None
                    return str(row[col]).strip()

                name = get_val("name")
                role = get_val("role")
                country = get_val("country")
                base_price_str = get_val("base_price")

                if not name or not role or not country or not base_price_str:
                    errors.append(f"Row {row_idx + 1}: missing required fields")
                    continue

                try:
                    base_price = float(base_price_str.replace(",", "").replace("₹", "").replace("$", ""))
                except ValueError:
                    errors.append(f"Row {row_idx + 1}: invalid base_price '{base_price_str}'")
                    continue

                def int_or_zero(val):
                    if not val:
                        return 0
                    try:
                        return int(float(val.replace(",", "")))
                    except:
                        return 0

                def float_or_zero(val):
                    if not val:
                        return 0.0
                    try:
                        return float(val.replace(",", ""))
                    except:
                        return 0.0

                player = Player(
                    auction_id=auction_id,
                    name=name,
                    role=role.lower().strip(),
                    country=country,
                    base_price=base_price,
                    image_url=get_val("image_url") or None,
                    status="unsold",
                    matches=int_or_zero(get_val("matches")),
                    runs=int_or_zero(get_val("runs")),
                    wickets=int_or_zero(get_val("wickets")),
                    batting_avg=float_or_zero(get_val("batting_avg")),
                    batting_sr=float_or_zero(get_val("batting_sr")),
                    bowling_avg=float_or_zero(get_val("bowling_avg")),
                    bowling_econ=float_or_zero(get_val("bowling_econ")),
                )
                db.add(player)
                players_created += 1
            except Exception as e:
                errors.append(f"Row {row_idx + 1}: {str(e)}")

        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

    return {
        "players_created": players_created,
        "errors": errors[:20],
        "total_rows": len(rows),
    }
